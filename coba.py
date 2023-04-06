import cv2
import numpy as np
import openpyxl
from datetime import datetime


cap = cv2.VideoCapture(0)
# cap = cv2.VideoCapture('pidio3.mp4')


net = cv2.dnn.readNetFromONNX("D:\\CODE\\python\\code\\data.onnx")
classes = ["tidak standar", "standar", "motor"]

# Tambahkan variabel counter untuk masing-masing class
counter = [0, 0]
line=600


# Load workbook dan worksheet
wb = openpyxl.load_workbook('D:/code/python/code/counter.xlsx')
ws = wb.active


# Tulis header pada worksheet
ws.cell(row=1, column=1, value='Knalpot Standar')
ws.cell(row=1, column=2, value='Knalpot Tidak Standar')




while True:
    img = cap.read()[1]
    if img is None:
        break
    img = cv2.resize(img, (640,640))
    blob = cv2.dnn.blobFromImage(img,scalefactor= 1/255,size=(640,640),mean=[0,0,0],swapRB= True, crop= False)
    net.setInput(blob)
    detections = net.forward()[0]
    current_counter = [0] * len(classes)
  
    classes_ids = []
    confidences = []
    boxes = []
    rows = detections.shape[0]

    img_width, img_height = img.shape[1], img.shape[0]
    x_scale = img_width/640
    y_scale = img_height/640

    for i in range(rows):
        row = detections[i]
        confidence = row[4]           
        if confidence > 0.2:
            classes_score = row[5:]
            ind = np.argmax(classes_score)
            if classes_score[ind] > 0.25:
                classes_ids.append(ind)
                confidences.append(confidence)
                cx, cy, w, h = row[:4]
                x1 = int((cx- w/2)*x_scale)
                y1 = int((cy-h/2)*y_scale)
                width = int(w * x_scale)
                height = int(h * y_scale)
                box = np.array([x1,y1,width,height])
                boxes.append(box)

    # for class_id in set(classes_ids):
    #     if counter[class_id] == 0:
    #         counter[class_id] =+ 1

    indices = cv2.dnn.NMSBoxes(boxes,confidences,0.1,0.1)

    for i in indices:
        x1,y1,w,h = boxes[i]
        class_id = classes_ids[i]
        label = classes[class_id]
        conf = confidences[i]
        text = label + " ({:.2f})".format(conf)
        cv2.rectangle(img,(x1,y1),(x1+w,y1+h),(255,0,0),2)
        cv2.putText(img, text, (x1,y1-2),cv2.FONT_HERSHEY_COMPLEX, 0.7,(0,0,255),2)
        
        # tambahkan logika deteksi melewati garis
        if y1 < 400 and y1 + h < 400:
            current_counter[class_id] += 1
        
    for i, count in enumerate(counter):
        if i in classes_ids:
            counter[i] += current_counter[i]
        else:
            counter[i] = max(0, count)

    # tambahkan kode untuk menampilkan counter pada gambar
    for class_id, count in enumerate(counter):
        label = classes[class_id]
        text = "{}: {}".format(label, count)
        cv2.putText(img, text, (10, 30 + (30 * class_id)), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (00, 255), 2)
        cv2.imshow("EXHAUST",img)

    # Mendapatkan waktu saat ini
    now = datetime.now()

    # Format tanggal dan waktu menjadi string
    date_string = now.strftime("%Y-%m-%d")
    time_string = now.strftime("%H:%M:%S")

    
    # Tulis header pada worksheet dengan tanggal dan waktu
    ws.cell(row=1, column=1, value='Knalpot Standar')
    ws.cell(row=1, column=2, value='Knalpot Tidak Standar')
    ws.cell(row=1, column=3, value='Tanggal')
    ws.cell(row=1, column=4, value='Waktu')
    ws.cell(row=2, column=1, value=counter[1])
    ws.cell(row=2, column=2, value=counter[0])
    ws.cell(row=2, column=3, value=date_string)
    ws.cell(row=2, column=4, value=time_string)

    # Simpan workbook
    wb.save('D:/code/python/code/counter.xlsx')
        

    k = cv2.waitKey(0)
    if k == ord('q'):
        break
    elif k == ord('r'):
        counter = [0, 0]
